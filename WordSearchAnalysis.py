import openpyxl
def Occurence(s, corpus):
	print(s)
	print(corpus)
	file = open(corpus, 'r')
	count = 0 
	for line in file:
		for word in (line.split()):
			if(word.lower() == s.lower()):
				count = count+1
				
	print(count)
	return count

wb = openpyxl.load_workbook('WordSearchResults.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# WS_Emotion = []
# RS_English = []
# DS_English = []
# DS_Hindi = []
# RS_Hindi = []
i=3
while(i<=205):
	D_Corpus_English = 0
	D_Corpus_Hindi = 0
	F_Corpus_English = 0
	F_Corpus_Hindi = 0
	English_word = ""
	Hindi_word = ""
	D_Corpus = "Dalit_Collated_Cleaned.txt"
	F_Corpus = "Final_Fem_Collated_Clean.txt"
	while(English_word != "Total" or Hindi_word !="Total"):
		English_word = sheet.cell(i,1).value
		Hindi_word = sheet.cell(i,4).value
		if(English_word is not None and (English_word != "Total")):
			#RS_English.append(s.encode('utf-8'))
			D_n = Occurence(English_word.encode('utf-8'), D_Corpus)
			#print(D_n)
			F_n = Occurence(English_word.encode('utf-8'), F_Corpus)
			#print(D_n)
			sheet.cell(i,2).value = D_n
			sheet.cell(i,3).value = F_n
			D_Corpus_English = D_Corpus_English + D_n
			F_Corpus_English = F_Corpus_English + F_n
		if(Hindi_word is not None and (Hindi_word != "Total")):
			#RS_English.append(s.encode('utf-8'))
			D_n = Occurence(Hindi_word.encode('utf-8'), D_Corpus)
			F_n = Occurence(Hindi_word.encode('utf-8'), F_Corpus)
			sheet.cell(i,5).value = D_n
			sheet.cell(i,6).value = F_n
			D_Corpus_Hindi = D_Corpus_Hindi + D_n
			F_Corpus_Hindi = F_Corpus_Hindi + F_n
		i=i+1
	sheet.cell(i-1,2).value = D_Corpus_English
	sheet.cell(i-1,3).value = F_Corpus_English
	sheet.cell(i-1,5).value = D_Corpus_Hindi
	sheet.cell(i-1,6).value = F_Corpus_Hindi
	D_Total_Occurence = D_Corpus_Hindi+D_Corpus_English
	F_Total_Occurence = F_Corpus_Hindi+F_Corpus_English
	if(D_Total_Occurence == 0):
		D_ratio = "NA"
	else:
		D_ratio = (float)(D_Corpus_Hindi)/(float)(D_Total_Occurence)*100
	if(F_Total_Occurence == 0):
		F_ratio = "NA"
	else:
		F_ratio = (float)(F_Corpus_Hindi)/(float)(F_Total_Occurence)*100
	sheet.cell(i-1,7).value = D_ratio
	sheet.cell(i-1,8).value = F_ratio
	i=i+1 #skip blank line
	






# while(i<8):	
# 	s = ""
# 	while(s != "z"):
# 		s = sheet.cell(i1,1).value
# 		if(s is not None and (s != "z")):
# 			RS_English.append(s.encode('utf-8'))
# 			D_n = Occurence(s.encode('utf-8'), D_Corpus)
# 			F_n = Occurence(s.encode('utf-8'), F_Corpus)
# 		i1=i1+1
# 	print RS_English

# 	s = ""
# 	while(s != "z"):
# 		s = sheet.cell(i2,2).value
# 		if(s is not None and (s != "z")):
# 			#DS_English.append(s.encode('utf-8'))
# 			#file.write(s.encode('utf-8'))
# 			D_n = Occurence(s.encode('utf-8'), D_Corpus)
# 			F_n = Occurence(s.encode('utf-8'), F_Corpus)
# 		i2=i2+1
# 	#print(DS_English)

# 	s = ""
# 	while(s != "z"):
# 		s = sheet.cell(i3,3).value
# 		if(s is not None and (s != "z")):
# 			#DS_Hindi.append(s.encode('utf-8'))
# 			D_n = Occurence(s.encode('utf-8'), D_Corpus)
# 			F_n = Occurence(s.encode('utf-8'), F_Corpus)
# 		i3=i3+1
# 	#print(DS_Hindi)
# 	s = ""
# 	while(s != "z"):
# 		s = sheet.cell(i4,2).value
# 		if(s is not None and (s != "z")):
# 			#RS_Hindi.append(s.encode('utf-8'))
# 			D_n = Occurence(s.encode('utf-8'), D_Corpus)
# 			F_n = Occurence(s.encode('utf-8'), F_Corpus)
# 		i4=i4+1
# 	#print(RS_Hindi)	
# 	intermediate = []
# 	intermediate.append(RS_English)

# 	#for elem in intermediate:
# 	#	print(elem)
# 	#WS_Emotion.append([RS_English, DS_English, DS_Hindi, RS_Hindi])
# 	WS_Emotion.append(intermediate)
# 	# del intermediate[:]
# 	# del RS_English[:]
# 	# del DS_English[:]
# 	# del DS_Hindi[:]
# 	# del RS_Hindi[:]
# 	i=i+1
wb.save("WordSearchResults.xlsx")
# for elem in WS_Emotion:
# 	print(elem)