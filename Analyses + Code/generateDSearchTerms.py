import openpyxl
wb = openpyxl.load_workbook('Keywords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet.cell(1, 1).value)
RS_D_Trigger = []
for i in range(2,52):
	RS_D_Trigger.append(sheet.cell(i,1).value)
#print(RS_D_Trigger)

RS_D_NonTrigger = []
for i in range(2,90):
	RS_D_NonTrigger.append(sheet.cell(i,4).value)
#print(RS_D_NonTrigger)

DS_D_Trigger = []
for i in range(2,45):
	DS_D_Trigger.append(sheet.cell(i,7).value)
#print(DS_D_Trigger)

DS_D_NonTrigger = []
for i in range(2,85):
	DS_D_NonTrigger.append(sheet.cell(i,10).value)
#print(DS_D_NonTrigger)
count_terms=0
D_Keyword = RS_D_Trigger + RS_D_NonTrigger + DS_D_Trigger + DS_D_NonTrigger
file = open("D_SearchTerms.txt", "w")
for first_key in RS_D_Trigger:
	for second_key in D_Keyword:
		if (first_key!=second_key):
			search_term = first_key+" "+second_key+ "\n"
			count_terms=count_terms+1
			file.write(search_term.encode('utf-8'))
D_Keyword = RS_D_NonTrigger + DS_D_Trigger + DS_D_NonTrigger
for first_key in DS_D_Trigger:
	for second_key in D_Keyword:
		if (first_key!=second_key):
			search_term = first_key+" "+second_key+ "\n"
			count_terms=count_terms+1
			file.write(search_term.encode('utf-8'))
print(count_terms)


#file.write(DS_D_Trigger[0].encode('utf-8'))

#for i in range(2, 52):