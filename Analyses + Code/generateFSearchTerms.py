import openpyxl
wb = openpyxl.load_workbook('F_Keywords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet.cell(1, 1).value)
RS_F_Trigger = []
for i in range(2,52):
	RS_F_Trigger.append(sheet.cell(i,1).value)
#print(RS_D_Trigger)

RS_F_NonTrigger = []
for i in range(2,98):
	RS_F_NonTrigger.append(sheet.cell(i,4).value)
#print(RS_D_NonTrigger)

DS_F_Trigger = []
for i in range(2,44):
	DS_F_Trigger.append(sheet.cell(i,7).value)
#print(DS_D_Trigger)

DS_F_NonTrigger = []
for i in range(2,97):
	DS_F_NonTrigger.append(sheet.cell(i,10).value)
#print(DS_D_NonTrigger)
count_terms=0
F_Keyword = RS_F_Trigger + RS_F_NonTrigger + DS_F_Trigger + DS_F_NonTrigger
file = open("F_SearchTerms.txt", "w")
for first_key in RS_F_Trigger:
	for second_key in F_Keyword:
		if (first_key!=second_key):
			search_term = first_key+" "+second_key+ "\n"
			count_terms=count_terms+1
			file.write(search_term.encode('utf-8'))
F_Keyword = RS_F_NonTrigger + DS_F_Trigger + DS_F_NonTrigger
for first_key in DS_F_Trigger:
	for second_key in F_Keyword:
		if (first_key!=second_key):
			search_term = first_key+" "+second_key+ "\n"
			count_terms=count_terms+1
			file.write(search_term.encode('utf-8'))
print(count_terms)


#file.write(DS_D_Trigger[0].encode('utf-8'))

#for i in range(2, 52):